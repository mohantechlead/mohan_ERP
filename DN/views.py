from django.shortcuts import get_object_or_404, render, redirect
from datetime import datetime
from django.core import serializers
from django.db.models import Sum, Case, When, F, Value, IntegerField
from django.contrib.postgres.aggregates import ArrayAgg
from django.db.models.functions import ExtractMonth,ExtractWeek
from .serializers import *
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.http import JsonResponse
from django.contrib import messages
from .forms import *
from FGRN.models import finished_goods
from MR.models import inventory
from .models import *
from django.forms import formset_factory
from django.contrib.auth.decorators import login_required
import plotly.graph_objs as go
from plotly.offline import plot
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment
from django.core.mail import send_mail
from django.core.serializers.json import DjangoJSONEncoder
import json

@api_view(['GET','POST'])
def order_list(request):
    if request.method == 'GET':
        my_orders = orders.objects.all()
        serializer =  OrderSerializer(my_orders,many=True)
        return JsonResponse({"Orders":serializer.data})

    if request.method == 'POST':
        serializer = OrderSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)


@api_view(['GET','POST'])

def delivery_list(request):
    if request.method == 'GET':
        my_deliveries = orders.objects.all()
        serializer =  DeliverySerializer(my_deliveries,many=True)
        return JsonResponse({"Deliveries":serializer.data})

    if request.method == 'POST':
        serializer = DeliverySerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status = status.HTTP_201_CREATED)

@login_required(login_url="login_user")
def input_delivery(request):
    my_orders = orders.objects.all()
    truck_number = sorted(set(delivery.objects.values_list('truck_number', flat=True)))
    driver_name = sorted(set(delivery.objects.values_list('driver_name', flat=True)))
    customer = sorted(set(Customer.objects.values_list('company', flat=True)))

    if request.method == 'POST':
        form = DeliveryForm(request.POST)
        delivery_number = request.POST['delivery_number']

        if form.errors:
            print(form.errors)
        
        if form.is_valid():
                serial_no = form.cleaned_data['serial_no']
                order = orders.objects.get(serial_no = serial_no.serial_no)       
                form.save()
                return redirect('input_delivery')
        
        else:
            print(form.data,"nval")
            errors = dict(form.errors.items())
            print(errors,"errors")
            return JsonResponse({'form_errors': errors}, status=400)
            
        
   
    form = DeliveryForm()
    formset = formset_factory(DeliverItemForm, extra= 1)
    formset = formset(prefix="items")

    return render(request, 'input_delivery.html', {'form': form,
                                                   'my_orders':my_orders, 
                                                   'formset':formset, 
                                                   'truck_number': truck_number, 
                                                   'my_customer': customer, 
                                                   'driver_name': driver_name,})

@login_required(login_url="login_user")
def input_delivery_items(request):
    
    if request.method == 'POST':
        print(request.POST)
        formset = formset_factory(DeliverItemForm, extra=1 , min_num= 1)
        formset = formset(request.POST or None, prefix="items")

        if formset.errors:
            print(formset.errors)

        non_empty_forms = [form for form in formset if form.cleaned_data.get('description')]
        delivery_no = request.POST.get('delivery_number')
        print(delivery_no,"delivery_no")
        if non_empty_forms:
            if formset.is_valid():
                try:
                    # Retrieve the delivery instance
                    Delivery_instance = delivery.objects.get(delivery_number=delivery_no)

                    for form in non_empty_forms:
                        form.instance.delivery_number = Delivery_instance

                        selected_item = form.cleaned_data['description']
                        # selected_item_description = selected_item.item_name  # Assuming item_name is the field
                        
                    
                        form.save()
            
                        Delivery_instance.save()

                    related_delivery = Delivery_instance

                    # Retrieve the serial_no from the delivery instance
                    delivery_serial_no = related_delivery.serial_no
                    print(f"Serial No: {delivery_serial_no}")

                    # Retrieve the delivery_items queryset filtered by the serial_no
                    delivery_items_queryset = delivery_items.objects.filter(delivery_number__serial_no=delivery_serial_no)

                    # Calculate the aggregated data grouped by serial_no and description
                    aggregated_data = (
                        delivery_items_queryset
                        .values('delivery_number__serial_no', 'description')  # Group by serial_no and description
                        .annotate(total_quantity=Sum('quantity'))  # Sum the quantities
                    )


                    # Display or process the aggregated results
                    for data in aggregated_data:
                        print(f"Serial No: {data['delivery_number__serial_no']}, Description: {data['description']}, Total Quantity: {data['total_quantity']}")

                    # Example: Process each item for further logic
                    for data in aggregated_data:
                        serial_no = data['delivery_number__serial_no']
                        description = data['description']
                        total_quantity = data['total_quantity']

                        # Find the matching order item in orders_items (if needed)
                        order_item = orders_items.objects.filter(serial_no=serial_no, description=description).first()
                        if order_item:
                            print(f"Before Update: Order Item {order_item.description}, Quantity: {order_item.quantity}")
                            
                            # Subtract the total quantity from the order item
                            order_item.quantity -= total_quantity
                                                        
                            print(f"After Update: Order Item {order_item.description}, Quantity: {order_item.quantity}")

                            if order_item.quantity < 0:
                                error_message = 'Over Delivery'
                                subject = 'Over Delivery Notification'
                                message = f'The order with Order ID {serial_no} has been over delivered by {order_item.quantity} on delivery number {delivery_no}'
                                from_email = 'tech@mohanplc.com'
                                recipient_list = ['tibarek90@gmail.com']

                                send_mail(subject, message, from_email, recipient_list, fail_silently= False)

                        else:
                            print(f"No matching order item found for Serial No: {serial_no} and Description: {description}")

                except delivery.DoesNotExist:
                    print(f"Delivery with Delivery_no {delivery_no} does not exist.")
                    return JsonResponse({'error': 'Invalid Delivery_no'}, status=400)
                
                except orders_items.DoesNotExist:
                    print(f"Order item with Serial_no {delivery_serial_no} does not exist.")
                    return JsonResponse({'error': 'Invalid Serial_no'}, status=400)
                                               
            else:
                print(formset.data,"nval")
                errors = dict(formset.errors.items())
                return JsonResponse({'form_errors': errors}, status=400)
        
            order_form = DeliveryForm(prefix="orders")
            formset = formset_factory(DeliverItemForm, extra=1)
            formset = formset(prefix="items")

            context = {
                'order_form': order_form,
                'formset': formset,
            }
            return render(request, 'input_delivery.html', context)
    else:
       
        formset = formset_factory(DeliverItemForm, extra=1)
        formset = formset(prefix="items")

    context = {
        'formset': formset,
    }
    return render(request, 'input_delivery.html', context)

@login_required(login_url="login_user")
#To display the delivery notes for a specific order number
def deliveries(request):
  my_orders = orders.objects.all()
  deliveries = []
  for order in my_orders:
    my_delivery = delivery.objects.filter(order_number=my_orders.order_number).first()
    if my_delivery:
      deliveries.append(my_delivery)

  context = {
    'orders': orders,
    'deliveries': deliveries,
  }

  return render(request, 'deliveries.html', context)
@login_required(login_url="login_user")
#for creating Orders
def input_orders(request):
    customer = Customer.objects.values_list('company', flat=True)
    
    
    if request.method == 'POST':
        form = OrderForm(request.POST)

        if form.errors:
            print(form.errors)

        if form.is_valid():
            order_number = form.cleaned_data['serial_no']
            form.save()
            return redirect('input_orders')
        
        else:
            print(form.data,"nval")
            errors = dict(form.errors.items())
            print(errors,"errors")
            return JsonResponse({'form_errors': errors}, status=400)
        
    my_goods = finished_goods.objects.all()
    form = OrderForm()
    formset = formset_factory(OrderItemForm, extra= 1)
    formset = formset(prefix="items")
    context = {'form': form ,'my_goods': my_goods, 'formset': formset, 'my_customer':customer}
    return render(request, 'input_orders.html', context)

@login_required(login_url="login_user")
def input_orders_items(request):
    if request.method == 'POST':
        print(request.POST)
        formset = formset_factory(OrderItemForm, extra=1 , min_num= 1)
        formset = formset(request.POST or None, prefix="items")

        if formset.errors:
            print(formset.errors)

        non_empty_forms = [form for form in formset if form.cleaned_data.get('description')]
        order_number = request.POST.get('serial_no')
        print(order_number,"order_no")
        if non_empty_forms:
            if formset.is_valid():
                try:
                    Order_instance = orders.objects.get(serial_no = order_number)
                    vat_amount = 0.0
                    final_price  = 0.0
                    before_vat = 0.0
                    final_unit = 0.0
                    for form in non_empty_forms:
                        form.instance.serial_no = Order_instance
                    
                        quantity = form.cleaned_data['quantity']
                        no_of_unit = form.cleaned_data['no_of_unit']
                        total_price = form.cleaned_data['total_price']
                        unit_price = form.cleaned_data['unit_price']
                
                        final_unit += no_of_unit
                        vat_amount += unit_price* quantity * 0.15
                        Order_instance.vat_amount = vat_amount
                        final_price += total_price + vat_amount
                        Order_instance.final_price = final_price
                        before_vat += total_price  
                        Order_instance.before_vat = before_vat
                        
                        
                        form.instance.remaining_quantity = quantity
                        form.instance.remaining_unit = no_of_unit
                        form.save()
            
                        Order_instance.save()
                                    
                except orders.DoesNotExist:
                    print(f"Order with Order_no {order_number} does not exist.")
                    return JsonResponse({'error': 'Invalid Order_no'}, status=400)
                    

                    
                    
            else:
                print(formset.data,"nval")
                errors = dict(formset.errors.items())
                return JsonResponse({'form_errors': errors}, status=400)
        
            order_form = OrderForm(prefix="orders")
            formset = formset_factory(OrderItemForm, extra=1)
            formset = formset(prefix="items")

            context = {
                'order_form': order_form,
                'formset': formset,
                # 'message':success_message,
            }
            return render(request, 'input_orders.html', context)
    else:
       
        formset = formset_factory(OrderItemForm, extra=1)
        formset = formset(prefix="items")

    context = {
        'formset': formset,
    }
    return render(request, 'input_orders.html', context)


@login_required(login_url="login_user")
def display_orders(request):
 
    the_order = orders.objects.all().order_by('serial_no')
    the_delivery = delivery.objects.all().order_by('delivery_number')

    orders_data = []

    for order in the_order:
        items = orders_items.objects.filter(serial_no=order.serial_no)
        dn = delivery.objects.filter(serial_no=order.serial_no)
        delivery_items_list = delivery_items.objects.filter(
                delivery_number__in=dn
            )
        
        for item in items:
            ordered_quantity = item.quantity
       
            delivered_quantity = delivery_items_list.filter(
                    description=item.description
                ).aggregate(total_delivered=Sum('quantity'))['total_delivered'] or 0
            
            remaining_quantity = ordered_quantity - delivered_quantity
            item.remaining_quantity = remaining_quantity

            print(item.remaining_quantity)
            item.save()
            
        order_data = {
                'serial_no': order.serial_no,
                'before_vat': order.before_vat,
                'invoice':order.invoice,
                'remaining_quantity': remaining_quantity,
                'date': order.date,
                'final_price': order.final_price,
                'order_item': items,
                'customer_name': order.customer_name 
            }
        orders_data.append(order_data)

    print(the_delivery)

    context = {
        'my_order': the_order,
        'orders_data': orders_data 
    }

    return render(request,'display_orders.html', context)

@login_required(login_url="login_user")
def display_single_fgrn(request):
    if request.method == 'GET':
        serial_no = request.GET['serial_no']
        
        try:
            fgrns = orders.objects.get(serial_no=serial_no)
            fgrn_items = orders_items.objects.all()
            fgrn_items = fgrn_items.filter(serial_no=serial_no)
            print(fgrn_items)

            if fgrn_items.exists():
                print(fgrn_items,"yes")
                context = {
                            'fgrn_item': fgrn_items,
                            'my_fgrn': fgrns,
                        }
                return render(request, 'display_single_fgrn.html', context)
        
        except orders.DoesNotExist:
                fgrns = None 
       
        print("no")
        
        context = {
                        'my_fgrn': fgrns,
                    }
    return render(request, 'display_single_fgrn.html')

@login_required(login_url="login_user")
def display_remaining(request):
    my_orders = orders.objects.filter(remaining__gt=0)
    return render(request, 'display_remaining.html', {'my_orders': my_orders})

@login_required(login_url="login_user")  
def display_single_order(request, serial_no):
    if request.method == 'GET':
        try:
            # Fetch the order with the given serial number
            fgrns = orders.objects.get(serial_no=serial_no)

            # Fetch the items for that order
            fgrn_items = orders_items.objects.filter(serial_no=serial_no)

            if fgrn_items.exists():
                context = {
                    'fgrn_item': fgrn_items,
                    'my_fgrn': fgrns,
                }
                return render(request, 'display_single_order.html', context)
            else:
                # If no items are found, return an empty list in the context
                context = {
                    'fgrn_item': [],
                    'my_fgrn': fgrns,
                }
                return render(request, 'display_single_order.html', context)
        
        except orders.DoesNotExist:
            # If the order does not exist, return None for 'my_fgrn' and an empty list for 'fgrn_item'
            context = {
                'fgrn_item': [],
                'my_fgrn': None,
            }
            return render(request, 'display_single_order.html', context)

       
        
    
@login_required(login_url="login_user")  
def display_single_delivery(request, delivery_number):
    if request.method == 'GET':
        try:
            # Fetch the delivery record with the given delivery number
            fgrns = delivery.objects.get(delivery_number=delivery_number)

            # Fetch the items for that delivery
            fgrn_items = delivery_items.objects.filter(delivery_number=delivery_number)

            if fgrn_items.exists():
                print(fgrn_items, "yes")
                context = {
                    'fgrn_item': fgrn_items,
                    'my_fgrn': fgrns,
                }
            else:
                # If no items are found, include an empty list in the context
                context = {
                    'fgrn_item': [],
                    'my_fgrn': fgrns,
                }
            return render(request, 'display_single_delivery.html', context)
        
        except delivery.DoesNotExist:
            # If the delivery does not exist
            print("Delivery not found")
            context = {
                'fgrn_item': [],
                'my_fgrn': None,
            }
            return render(request, 'display_single_delivery.html', context)

@login_required(login_url="login_user")
def display_delivery(request):
    my_delivery = delivery.objects.all()
    # Sort and filter based on user input
    sort_param = request.GET.get('sort')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Filter deliveries by date range
    if start_date and end_date:
        my_delivery = my_delivery.filter(delivery_date__range=[start_date, end_date])
        my_delivery = my_delivery.order_by('delivery_date')
    elif sort_param == 'date':
        my_delivery = my_delivery.order_by('delivery_date')
    elif sort_param == 'customer_name':
        my_delivery = my_delivery.order_by('serial_no__customer_name')  # Sorting by related orders' customer_name
    elif sort_param == 'delivery_number':
        my_delivery = my_delivery.order_by('delivery_number')

    # Prepare delivery data
    deliveries_data = []

    for deliveries in my_delivery:
        # Fetch related delivery items
        items = delivery_items.objects.filter(delivery_number=deliveries.delivery_number)
        
        delivery_data = {
            'delivery_number': deliveries.delivery_number,
            'serial_no': deliveries.serial_no,
            'delivery_date': deliveries.delivery_date,
            'total_quantity': deliveries.total_quantity,
            'total_bags': deliveries.total_bags,
            'truck_number': deliveries.truck_number,
            'driver_name': deliveries.driver_name,
            'recipient_name': deliveries.recipient_name,
            'items': items,  # Add the related items to each delivery
        }
        deliveries_data.append(delivery_data)

    # Render the deliveries and their items to the template
    return render(request, 'display_delivery.html', {
        'my_delivery': my_delivery,
        'deliveries_data': deliveries_data,
    })

@login_required(login_url="login_user")
def search_customer(request):
    if request.method == 'GET':
        customer_name = request.GET['customer_name']
        #my_order = get_object_or_404(orders, customer_name=customer_name)
        my_order = orders.objects.filter(customer_name__icontains= customer_name)
        
        if my_order.exists():
            context = {
                        'my_order': my_order,
                    }
            return render(request, 'customer_details.html', context)
            
        else:
            print("none")
            return render(request, 'customer_details.html', context)

    else:
        context = {
                        'my_order': my_order,
                    }
        return render(request, 'customer_details.html', context)
# Create your views here.

@login_required(login_url="login_user")
def customer_date(request):
    customer_name = request.GET.get('customer_name')
    my_order = orders.objects.filter(customer_name__icontains= customer_name)
    my_delivery = delivery.objects.filter('serial_no__customer_name')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Filter orders within the specified date range
    if start_date and end_date:
        my_delivery = my_delivery.filter(delivery_date__range=[start_date, end_date])
        my_delivery = my_delivery.order_by('delivery_date')
    context = {
                    'my_orders': my_delivery,
                }
    if request.method == 'GET':
        customer_name = request.GET['customer_name']
        my_order = orders.objects.filter(customer_name__icontains= customer_name)
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Filter orders within the specified date range
        if start_date and end_date:
            my_order = my_order.filter(delivery_date__range=[start_date, end_date])
            my_order = my_order.order_by('delivery_date')
        context = {
                        'my_order': my_order,
                    }
    else:
        context = {
                        'my_order': my_order,
                    }
        return render(request, 'customer_details.html', context)

@login_required(login_url="login_user")
def search_orders(request):
    print("in")
    if request.method == 'GET':
        serial_no = request.GET['serial_no']

        # Get the order by serial number
        my_order = get_object_or_404(orders, serial_no=serial_no)
        # If the order exists,
        if my_order:
            print(my_order)
            print('yes')
            context = {
                        'my_order': my_order,
                    }
            return render(request, 'single_order.html', context)
            
        else:
            print("none")
            return render(request, 'single_order.html', context)

    else:
        context = {
                        'my_order': my_order,
                    }
        return render(request, 'display_orders.html')
# Create your views here.
@login_required(login_url="login_user")
def search_delivery(request):
    if request.method == 'GET':
        serial_no = request.GET['serial_no']
        my_order = get_object_or_404(orders, serial_no=serial_no)
        delivery_qs = delivery.objects.all()
        deliveries = delivery_qs.filter(serial_no=serial_no)
        if deliveries.exists():
            context = {
                        'deliveries': deliveries,
                        'my_order': my_order,
                    }
            return render(request, 'single_delivery.html', context)
            
        else:
            print("none")
            return render(request, 'single_delivery.html', context)
    
    
    else:
        context = {
                        'deliveries': deliveries,
                    }
        return render(request, 'single_delivery.html', context)

@login_required(login_url="login_user")
def search_customer_delivery(request):
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if request.method == 'GET':
        customer_name = request.GET.get('customer_name')
        my_order = orders.objects.filter(customer_name__icontains=customer_name)
        delivery_qs = delivery.objects.filter(serial_no__customer_name=customer_name)

        if start_date and end_date:
            # Parse the date strings to datetime objects
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Apply date filtering
            delivery_qs = delivery_qs.filter(delivery_date__range=[start_date, end_date])

        # Order the filtered deliveries
        my_delivery = delivery_qs.order_by('delivery_date')

        context = {
            'deliveries': my_delivery,
            'my_order': my_order,
            'customer_name': customer_name,
        }
        return render(request, 'customer_delivery.html', context)
    else:
        context = {
            'deliveries': [],
        }
        return render(request, 'customer_delivery.html', context)

@login_required(login_url="login_user")
def dashboard_with_pivot(request):
    return render(request, 'analytics_dashboard.html', {})

@login_required(login_url="login_user")
def pivot_data(request):
    dataset = orders.objects.all()
    data = serializers.serialize('json', dataset)
    return JsonResponse(data, safe=False)

@login_required(login_url="login_user")
def dashboards(request):
    # Fetch item names and their corresponding quantities
    qs = inventory_order_items.objects.all()
    
    # Prepare data for Chart.js
    labels = [item.item_name for item in qs]  # Extract item names
    data = [item.total_quantity for item in qs]  # Extract quantity (or any relevant field)

    # Convert to JSON for use in JavaScript
    context = {
        "labels": json.dumps(labels, cls=DjangoJSONEncoder),
        "data": json.dumps(data, cls=DjangoJSONEncoder),
    }

    return render(request, "dashboard.html", context)




@login_required(login_url="login_user")
def customer_table(request):
    my_customers = orders.objects.all()
  
    my_order = orders.objects.all()
    # Join the two tables on the serial_no column
    sort_param = request.GET.get('sort')

    # Sort the orders based on the selected parameter
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Filter orders within the specified date range
    if start_date and end_date:
        my_order = my_order.filter(date__range=[start_date, end_date])
        my_order = my_order.order_by('date')
   
    
    customer_data = my_order.values('customer_name').annotate(
        total_sales_kgs=Sum(
            Case(
                When(measurement='kgs', then=F('order_quantity')),
                default=Value(0),
                output_field=IntegerField(),
            )
        ),
        total_sales_pairs=Sum(
        Case(
            When(measurement='pairs', then=F('order_quantity')),
            default=Value(0),
            output_field=IntegerField(),
        )
        ),
        total_sales_pcs=Sum(
            Case(
                When(measurement='pcs', then=F('order_quantity')),
                default=Value(0),
                output_field=IntegerField(),
            )
        ),
    
        total_quantity=Sum('order_quantity'),
        item_list=ArrayAgg('description', distinct=True)  # Assuming 'item' is the name of the field for items
    )
    
    json_data = {
        'data': [
            {
                'customer_name': entry['customer_name'],
                'total_sales_kgs': entry['total_sales_kgs'],
                'total_quantity': entry['total_quantity'],
                'item_list': entry['item_list'],
                'total_sales_pairs': entry['total_sales_pairs'],
                'total_sales_pcs': entry['total_sales_pcs']
            }
            for entry in customer_data
        ]
    }
    return render(request, 'dashboard_tables.html', {'json_data': json_data, })

@login_required(login_url="login_user")
def item_table(request):
    item_data = orders.objects.filter(measurement='pairs').values('description').annotate(
        total_sales=Sum('total_price'),
        total_quantity=Sum('order_quantity') # Assuming 'item' is the name of the field for items
    )
    
    json_data = {
        'data': [
            {
                'description': entry['description'],
                'total_sales': entry['total_sales'],
                'total_quantity': entry['total_quantity']
            }
            for entry in item_data
        ]
    }
    return render(request, 'item_table.html', {'json_data': json_data})

@login_required(login_url="login_user")
def sales_contract(request):
    if request.method == 'GET':
    
        serial_no = request.GET['serial_no']
        print(serial_no,"seirla")
        my_order = get_object_or_404(orders, serial_no=serial_no)
        return render(request,'sales_contract.html',{'my_order':my_order})

    else:
        
        return render(request,'sales_contract.html')
    
def create_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('customer_list')  # Redirect to a list of customers (or wherever you prefer)
    else:
        form = CustomerForm()
    
    return render(request, 'create_customer.html', {'form': form})

def customer_list(request):
    customers = Customer.objects.all()
    return render(request, 'customer_list.html', {'customers': customers})

def order_chart(request):
    # Aggregating total quantity for each customer using `orders_items`
    customer_data = (
        orders_items.objects
        .values('serial_no__customer_name')  # Access customer name through `serial_no` (ForeignKey)
        .annotate(total_quantity=Sum('quantity'))  # Sum up the quantities
        .order_by('-total_quantity')  # Sort by quantity, descending
    )

    # Prepare the labels (customer names) and data (total quantity)
    labels = [data['serial_no__customer_name'] for data in customer_data]
    data = [data['total_quantity'] for data in customer_data]

    # Pass data to the template
    context = {
        'labels': labels,
        'data': data
    }
    
    return render(request, 'order_chart.html', context)

@login_required(login_url="login_user")
def display_DN_items(request):

    item_quantities = delivery_items.objects.values('description').annotate(total_quantity=Sum('quantity'), total_no_of_unit=Sum('no_of_unit'))
  
    for item in item_quantities:
        inventory_DN_items.objects.update_or_create(
            item_name=item['description'],
            defaults={'total_quantity': item['total_quantity'],
                      'total_no_of_unit': item['total_no_of_unit']}
            
        )
        print(f"Name: {item['description']}, Total Quantity: {item['total_quantity']}, Total No of Unit: {item['total_no_of_unit']}")
    

    items = inventory_DN_items.objects.all().order_by('item_name')    
    print(items)
    context = {
        # 'total_quantity':item_quantities,
        'items':items,
        
    }

    return render(request,'display_DN_items.html',context)

@login_required(login_url="login_user")
def display_order_items(request):
    if request.method == 'POST':
        form = OrderInventoryForm(request.POST)

        if form.errors:
            print(form.errors)

        if form.is_valid():
            form.save()
            return redirect('display_order_items')
    
    form = OrderInventoryForm()

    item_quantities = orders_items.objects.values('description').annotate(total_quantity=Sum('quantity'), total_no_of_unit=Sum('no_of_unit'))
  
    for item in item_quantities:
        inventory_order_items.objects.update_or_create(
            item_name=item['description'],
            defaults={'total_quantity': item['total_quantity'],
                      'total_no_of_unit': item['total_no_of_unit']}
            
        )
        print(f"Name: {item['description']}, Total Quantity: {item['total_quantity']}, Total No of Unit: {item['total_no_of_unit']}")
    

    items = inventory_order_items.objects.all().order_by('item_name')    

    group_name = finished_goods.objects.all().order_by('item_name')
    print(items)
    context = {
        'group':group_name,
        'items':items,
        'form':form,
    }

    return render(request,'display_order_items.html',context)

def deliveries_excel(request):
    # Fetch all unique company names
    companies = orders.objects.values_list('customer_name', flat=True).distinct()

   # Prepare data grouped by company
    company_data = {}
    for company in companies:
        # Fetch orders for the company
        orders_queryset = orders.objects.filter(customer_name=company)

        data = []
        for o in orders_queryset:
            items = orders_items.objects.filter(serial_no=o.serial_no)
            deliveries = delivery.objects.filter(serial_no=o.serial_no)

            # Initialize data for each order item
            for item in items:
                # Start with the initial quantity for each item
                initial_balance = item.quantity

                order_data = {
                    'order': {
                        'serial_no': o.serial_no,
                        'invoice_no': o.invoice,
                        'unit_price': item.unit_price,
                        'total_price': item.total_price,
                        'item_name': item.description,
                        'date': o.date,
                        'quantity': item.quantity,
                    },
                    'deliveries': [],
                    'updated_balance': initial_balance,  # Start with the initial quantity
                }

                # Loop through each delivery and update the balance for the matching item
                for delivry in deliveries:
                    # Get the items for this delivery from the delivery_item model
                    delivery_items_set = delivery_items.objects.filter(delivery_number=delivry.delivery_number)

                    for del_item in delivery_items_set:
                        if del_item.description == item.description:  # Match by item name
                            order_data['deliveries'].append({
                                'delivery_no': delivry.delivery_number,
                                'delivery_date': delivry.delivery_date,
                                'item_name': del_item.description,
                                'unit_price':item.unit_price,
                                'total_price':item.total_price,
                                'del_quantity': del_item.quantity,
                            })

                            # Update the balance by subtracting the delivered quantity
                            order_data['updated_balance'] -= del_item.quantity

                # Append the order data to the company's list
                data.append(order_data)

        # Assign collected data to the company
        company_data[company] = data


        
    return render(request, 'deliveries_excel.html', {'company_data': company_data})


def generate_excel(request):
    # Fetch all unique company names
    companies = orders.objects.values_list('customer_name', flat=True).distinct()

    # Create a workbook
    wb = openpyxl.Workbook()

    # Loop through each company and create a new sheet for each one
    for company in companies:
        # Create a sheet for each company, named after the company
        ws = wb.create_sheet(title=company)

        # Fetch orders for the company
        orders_queryset = orders.objects.filter(customer_name=company)

        # Create the header row
        ws.append([
            "Serial No", "Date", "DN No", "Cash Sales Invoice", "Price", 
            "TOTAL", "Balance", "Debit", "Credit", "Balance"
        ])

        # Loop through the orders and add rows to the sheet
        for o in orders_queryset:
            items = orders_items.objects.filter(serial_no=o.serial_no)
            deliveries = delivery.objects.filter(serial_no=o.serial_no)

            # Loop through each item in the order
            for item in items:
                initial_balance = item.quantity  # Start with the initial quantity for the item

                # Create the order row for the item (Initial row with initial balance)
                row = [
                    o.serial_no, o.date, "", item.description, "", 
                    "", item.total_price, "", item.quantity, initial_balance
                ]
                ws.append(row)
                order_row_index = len(list(ws.rows))  # Track the row index of the order row

                # Keep track of the delivery numbers we've processed for this item
                processed_delivery_numbers = set()

                # Add delivery rows only for unique delivery numbers
                for delivry in deliveries:
                    delivery_items_set = delivery_items.objects.filter(delivery_number=delivry.delivery_number)

                    # Process the deliveries only if the item description matches and the delivery hasn't been processed yet
                    for del_item in delivery_items_set:
                        if del_item.description == item.description:  # Match by item name
                            # Skip this delivery if it has already been processed
                            if delivry.delivery_number not in processed_delivery_numbers:
                                # Update the initial balance (subtract the delivered quantity)
                                initial_balance -= del_item.quantity
                                
                                # Create the delivery row
                                delivery_row = [
                                    "", delivry.delivery_date, delivry.delivery_number, 
                                    o.invoice, item.unit_price, item.total_price, "", "", 
                                    del_item.quantity,  # Delivered quantity
                                ]
                                ws.append(delivery_row)

                                # Update the order row with the updated balance (same row index)
                                ws.cell(row=order_row_index, column=10, value=initial_balance)  # Update the Balance cell

                                # Mark this delivery number as processed
                                processed_delivery_numbers.add(delivry.delivery_number)

    # Remove the default sheet created by openpyxl if it exists
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create a response to serve the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=company_data.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

def get_order_items(request):
    serial_no = request.GET.get('serial_no')  # Get order number (serial_no) from the query params
    items = set(orders_items.objects.filter(serial_no=serial_no).values_list('description', flat=True))
    
    return JsonResponse({'items': list(items)})

@login_required(login_url="login_user")
def customer_detail(request, company):
    customer = get_object_or_404(Customer, company=company)
       
    context = {
        'customer': customer,
        
    }
    return render(request, 'customer_details.html', context)



